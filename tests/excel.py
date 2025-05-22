from openpyxl.reader.excel import load_workbook

wb = load_workbook('C:/Users/rudre/PycharmProjects/PyTest_DataDriven_Framework/testdata/test_data.xlsx')
sheet=wb["Register"]
#sheet = wb.get_sheet_by_name('Register')
data = []
for row in sheet.iter_rows(values_only=True):
    if 'Yes' in row:
        data.append(row)
print(data)


# for row in (1,sheet.max_row+1):
#     for col in (1,sheet.max_column+1):
#         if sheet.cell(row=row,column=col+1).value != "None":
#             a = sheet.cell(row=row,column=col+1).value
#             b = b + str(a) + ","
#     print(b)