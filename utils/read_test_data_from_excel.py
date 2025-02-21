from openpyxl import load_workbook


def read_excel_test_data(file_path, sheet_name):
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]
    data = []
    headers = [cell.value for cell in sheet[1]]
    for row in range(2, sheet.max_row + 1):
        row_data = {}
        for col, header in enumerate(headers):
            cell_value = sheet.cell(row=row, column=col + 1).value
            row_data[header.strip('/')] = cell_value
        data.append(row_data)
    return data



